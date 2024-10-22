from datetime import datetime, timedelta
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl.styles import PatternFill, Alignment
import numpy as np
from openpyxl.utils import get_column_letter
from colored_headers import headers_to_color
import psutil


def perform_vlookup(button_to_disable):
    try:
        # ------------------ SELECT CONTRACT FILE -------------------------------
        # Ask the user for the contract file
        contract_file = filedialog.askopenfilename(title="Select the contract file, where we need a vlookup",
                                                   initialdir="H:\Program_Testing_Exec\Sorting_Creation_Updated\testing_new_logic_5_13_24")

        # ------------------ LOAD DATAFRAMES ------------------------------------
        # Load 'Active Supplier Contracts' and 'Prev Contract' sheets
        active_supplier_df = pd.read_excel(contract_file, sheet_name='Active Supplier Contracts', header=1,dtype={'IPN': str})
        prev_contract_df = pd.read_excel(contract_file, sheet_name='Prev Contract', header=0, dtype={'IPN': str, 'PSoft Part': str})
        lost_items_df = pd.read_excel(contract_file, sheet_name='Lost Items', dtype={'IPN': str})
        awards_df = pd.read_excel(contract_file, sheet_name='Awards', dtype={'Award CPN': str})
        snd_df = pd.read_excel(contract_file, sheet_name='SND', dtype={'IPN': str})
        vpc_df = pd.read_excel(contract_file, sheet_name='VPC', dtype={'IPN': str})
        backlog_df = pd.read_excel(contract_file, sheet_name='Backlog', dtype={'IPN': str})
        sales_history_df = pd.read_excel(contract_file, sheet_name='Sales History', dtype={'IPN': str})
        running_file_df = pd.read_excel(contract_file, sheet_name='Price Increases',dtype={'Creation Part Number': str})

        print("Loaded 'Active Supplier Contracts' sheet with shape:", active_supplier_df.shape)
        print("Loaded 'Prev Contract' sheet with shape:", prev_contract_df.shape)
        # ---------------- End of Loading Sheets ---------------------------------

        # Ensure 'IPN' and 'Award CPN' columns are of string type and have leading zeros
        dataframes = [active_supplier_df, prev_contract_df, lost_items_df, awards_df, snd_df, vpc_df, backlog_df,
                      sales_history_df, running_file_df]
        for df in dataframes:
            if 'IPN' in df.columns:
                df['IPN'] = df['IPN'].astype(str).str.zfill(8)  # Convert to string and pad with leading zeros
            if 'Award CPN' in df.columns:
                df['Award CPN'] = df['Award CPN'].astype(str).str.zfill(
                    8)  # Convert to string and pad with leading zeros
        # Adding a value here, so making 6 to 8 and so on makes the difference of how many leading 0s there are in
        # the IPN

        # ------------------ PRELIMINARY DATA PREPARATION ----------------------------
        # Drop the 'LW PRICE' column if it exists in the 'Prev Contract' dataframe
        if 'LW PRICE' in prev_contract_df.columns:
            prev_contract_df.drop('LW PRICE', axis=1, inplace=True)

        # Rename the 'Price' column from 'Prev Contract' dataframe to 'LW PRICE'
        prev_contract_df.rename(columns={'Price': 'LW PRICE'}, inplace=True)

        # Rename 'MPN' in prev_contract_df to 'LW MPN'
        prev_contract_df.rename(columns={'MPN': 'LW MPN'}, inplace=True)

        # Merge 'LW MPN' from prev_contract_df into active_supplier_df as 'Prev Contract MPN'
        active_supplier_df = pd.merge(
            active_supplier_df,
            prev_contract_df[['IPN', 'LW MPN']],
            on='IPN',
            how='left'
        ).rename(columns={'LW MPN': 'Prev Contract MPN'})

        def update_mpn_match(df):
            df['MPN Match'] = np.where(df['MPN'] == df['Prev Contract MPN'], 'Y', 'N')
            return df

        # Apply the function to active_supplier_df
        active_supplier_df = update_mpn_match(active_supplier_df)

        # ------------------ UPDATE AWARDS DETAILS IN ACTIVE SUPPLIER DATAFRAME ------------------
        for idx, row in active_supplier_df.iterrows():
            ipn = row['IPN']  # we use the IPN value to go back and forth between the awards dataframe

            # Apply matching logic to find matching indices
            matching_indices = awards_df[awards_df['Award CPN'] == ipn].index

            if not matching_indices.empty:
                # Convert 'End Date' to datetime and then to the desired string format
                awards_df.loc[matching_indices, 'End Date'] = pd.to_datetime(
                    awards_df.loc[matching_indices, 'End Date'],
                    errors='coerce').dt.strftime('%m-%d-%Y')

                # Drop rows where 'End Date' conversion resulted in NaT
                awards_df = awards_df.dropna(subset=['End Date'])

                # Recheck matching indices after potential row drops
                matching_indices = awards_df[awards_df['Award CPN'] == ipn].index

                if not awards_df.loc[matching_indices].empty:
                    valid_end_dates = awards_df.loc[matching_indices]
                    latest_end_date = valid_end_dates['End Date'].max()
                    active_supplier_df.at[idx, 'PS Award Exp Date'] = latest_end_date

                    latest_price_row = valid_end_dates[valid_end_dates['End Date'] == latest_end_date]
                    if pd.notna(latest_price_row['Award Price'].iloc[0]):
                        active_supplier_df.at[idx, 'PS Award Price'] = latest_price_row['Award Price'].iloc[0]

                # Update 'PS Awd Cust ID' with the first non-NA 'Award Cust ID' found for matched IPN
                if matching_indices.any():
                    first_matched_cust_id = awards_df.loc[matching_indices, 'Award Cust ID'].dropna().iloc[0]
                    if pd.notna(first_matched_cust_id):
                        active_supplier_df.at[idx, 'PS Awd Cust ID'] = first_matched_cust_id

        # Convert 'PS Award Price' to numeric
        active_supplier_df['PS Award Price'] = pd.to_numeric(active_supplier_df['PS Award Price'], errors='coerce')
        awards_df['Award Price'] = pd.to_numeric(awards_df['Award Price'], errors='coerce')

        # Create a dictionary for 'Award CPN' and their 'Award Price' from awards_df
        award_price_mapping = awards_df.set_index('Award CPN')['Award Price'].to_dict()

        # Perform the comparison and update 'Price Match Award'
        active_supplier_df['Price Match Award'] = active_supplier_df.apply(
            lambda x: 'Y' if np.isclose(award_price_mapping.get(x['IPN'], np.nan), x['PS Award Price'],
                                        atol=1e-5) else 'N', axis=1
        )
        # ----------------------- End of Update Awards Detail in active dataframe -----------------------

        # ------------------ UPDATE 'CORP AWARD LOADED' STATUS ------------------
        # Normalize 'Award CPN' values from awards_df for a more accurate lookup (trim spaces, ensure consistent case)
        normalized_award_cpn_set = set(awards_df['Award CPN'].str.strip().str.lower())

        # Update 'Corp Award Loaded' based on the presence of a normalized 'IPN' in the normalized award CPN set
        # Use a different variable name inside the lambda to avoid shadowing
        active_supplier_df['Corp Awd Loaded'] = active_supplier_df['IPN'].str.strip().str.lower().apply(
            lambda x: 'Y' if x in normalized_award_cpn_set else 'N'
        )
        # ------------------ End of Corp Award loaded status --------------------

        # ---------------- Adding the merge algorithm to bring in columns from prev contract dataframe --------------
        active_supplier_df = active_supplier_df.merge(
            prev_contract_df[
                ['IPN', 'LW PRICE', 'PSoft Part',
                 "Contract Change", "count",
                 "SUM", "AVG", "DIFF", "PSID All Contract Prices Same?",
                 "90 DAY PI - NEW PRICE", "PI SENT DATE",
                 "DIFF Price Increase", "PI EFF DATE", "12 Month CPN Sales", "GP%", "Cost",
                 "Cost Note", "Quote#", "Cost Exp Date", "Cost MOQ",
                 "Review Note"]], on='IPN', how='left')
        # ---------------- End of Adding the merge algorithm to bring in columns from prev contract dataframe -------

        # ---------------- Formulas for the Sum Diff and Avg columns --------------------------------------
        # This will create a 'SUM' column based on the sum of 'Price' for each 'PSoft Part'
        active_supplier_df['SUM'] = active_supplier_df.groupby('PSoft Part')['Price'].transform('sum')

        # Assuming you already have a 'count' column calculated
        active_supplier_df['AVG'] = active_supplier_df['SUM'] / active_supplier_df['count']

        # Assuming 'Price' is the column you want to subtract from AVG to calculate DIFF
        active_supplier_df['DIFF'] = active_supplier_df['AVG'] - active_supplier_df['Price']
        # ---------------- End of Formulas for the Sum Diff and Avg columns --------------------------------------

        #  ------Calculate the counts for each 'PSoft Part' and display them with the value number we have ----------
        psoft_part_counts = active_supplier_df['PSoft Part'].value_counts()
        active_supplier_df['count'] = active_supplier_df['PSoft Part'].map(psoft_part_counts)

        active_supplier_df['IPN'] = active_supplier_df['IPN'].astype(str).str.strip()
        active_supplier_df['PSoft Part'] = active_supplier_df['PSoft Part'].astype(str).str.strip()
        awards_df.columns = awards_df.columns.str.strip()
        #  ------ End of Calculate the counts for each 'PSoft Part' and display them with the value number we have --

        # ------- This for loop is used for the next coming if statements and adjusting from other df ---------------
        for idx, row in active_supplier_df.iterrows():
            ipn = row['IPN']  # we use the IPN value to go back and forth between the awards dataframe
            psoft_part = row['PSoft Part']

            # ----------------- Bringing in MPN Column and then renaming it for comparison --------------------

            # ------------------ Running File logic to check IPN and bring in Unit Price New ------------------
            # This logic checks the IPN in the active dataframe and the creation part number in the running file and
            # brings in the Unit Price New in our 90-Day Pi Sent day column in our active workbook, it does this also
            # for all the data from the running file that we need in the active workbook, therefor it does this for
            # "90 DAY PI - NEW PRICE", "PI SENT DATE","PI EFF DATE",
            # Search for a matching IPN value in the running_file_df
            # Get the current year
            current_year = datetime.now().year

            # Search for a matching IPN value in the running_file_df for "90 DAY PI - NEW PRICE"
            matching_running_file = running_file_df[running_file_df['Creation Part Number'] == ipn]
            if not matching_running_file.empty and not pd.isna(matching_running_file.iloc[0, 11]):
                active_supplier_df.at[idx, '90 DAY PI - NEW PRICE'] = matching_running_file.iloc[0, 11]

            # For "PI SENT DATE" with the year check
            matching_running_file = running_file_df[running_file_df['Creation Part Number'] == ipn]
            if not matching_running_file.empty:
                pi_sent_date_raw = matching_running_file.iloc[0, 21]  # Assuming column V
                pi_sent_date = pd.to_datetime(pi_sent_date_raw, errors='coerce')  # Convert to datetime
                if pd.notnull(pi_sent_date) and pi_sent_date.year == current_year:
                    active_supplier_df.at[idx, 'PI SENT DATE'] = pi_sent_date

            # For "PI EFF DATE" with the year check
            matching_running_file = running_file_df[running_file_df['Creation Part Number'] == ipn]
            if not matching_running_file.empty:
                pi_eff_date_raw = matching_running_file.iloc[0, 4]  # Assuming column E
                pi_eff_date = pd.to_datetime(pi_eff_date_raw, errors='coerce')  # Convert to datetime
                if pd.notnull(pi_eff_date) and pi_eff_date.year == current_year:
                    active_supplier_df.at[idx, 'PI EFF DATE'] = pi_eff_date

            # Check if a match is found for price difference calculation
            if not matching_running_file.empty:
                # Extract the price from the active_supplier_df for the current IPN
                active_price = active_supplier_df.loc[idx, 'Price']
                running_price = matching_running_file.iloc[0, 11]  # Assuming column L for price

                # Check if both prices are not NaN and calculate the difference
                if not pd.isna(active_price) and not pd.isna(running_price):
                    price_difference = active_price - running_price
                    active_supplier_df.at[idx, 'DIFF Price Increase'] = price_difference

            # ------------------- Logic for updated the LW Cost and so on --------------------------------------
            # Assuming the IPN is in the first column for both dataframes
            # Iterate through the active_supplier_df to update each row based on matching IPN in prev_contract_df
            # Ensure IPN columns in both DataFrames are of the same type and prepared for matching
            # active_supplier_df['IPN'] = active_supplier_df['IPN'].astype(str).str.strip()
            # prev_contract_df['IPN'] = prev_contract_df['IPN'].astype(str).str.strip()

            # Reset 'Review Note' column to empty for user input and backup existing notes to 'LW Review Note'
            if 'Review Note' in active_supplier_df.columns:
                # Backup current review notes to 'LW Review Note'
                active_supplier_df['LW Review Note'] = active_supplier_df['Review Note']

            # Initialize 'Review Note' as an empty column
            active_supplier_df['Review Note'] = ''

            # ---------------- End of PRELIMINARY DATA PREPARATION ------------------------

            # Perform the merge operation to bring in the 'Cost' related columns from prev_contract_df
            # Here, we're creating a new DataFrame as a result of this merge to review the merge result before applying it back to active_supplier_df
            merged_df = pd.merge(
                active_supplier_df,
                prev_contract_df[['IPN', 'Cost', 'Cost Note', 'Quote#', 'Cost Exp Date', 'Review Note']],
                on='IPN',
                how='left',
                suffixes=('', '_prev')
            )

            # Now, update active_supplier_df with the merged data
            # This avoids direct row-by-row iteration and uses efficient pandas operations
            active_supplier_df['LW Cost'] = merged_df['Cost_prev']
            active_supplier_df['LW Cost Note'] = merged_df['Cost Note_prev']
            active_supplier_df['LW Quote#'] = merged_df['Quote#_prev']
            active_supplier_df['LW Cost Exp Date'] = merged_df['Cost Exp Date_prev']
            active_supplier_df['LW Review Note'] = merged_df[
                'Review Note_prev']  # Update from previously backed-up data

            # ------------------ BACKLOG VALUE MAPPING TO LOST ITEMS ------------------
            # Create a mapping from 'Backlog CPN' to 'Backlog Value' in backlog_df
            backlog_value_mapping = backlog_df.set_index('Backlog CPN')['Backlog Value'].to_dict()
            # Apply the mapping to 'IPN' column in lost_items_df to get 'Backlog Value'
            lost_items_df['Backlog Value'] = lost_items_df['IPN'].map(backlog_value_mapping)
            # Format the 'Backlog Value' column as currency with 2 decimal places
            lost_items_df['Backlog Value'] = lost_items_df['Backlog Value'].apply(
                lambda x: "" if pd.isnull(x) or x == 0 else "${:,.2f}".format(x))
            # ---------------------------------------------------------------------------

            # ------------------- 12 Month CPN Sales Column Logic ----------------------
            # Example column name: 'YourDateColumnName'
            date_column = 'Last Ship Date'  # Replace with your actual date column name
            # Convert the date column to datetime format
            sales_history_df[date_column] = pd.to_datetime(sales_history_df[date_column], errors='coerce')
            # Calculate the date for 12 months ago
            one_year_ago = datetime.now() - timedelta(days=365)
            # Filter sales_history_df to include only the last 12 months
            sales_history_df_filtered = sales_history_df[sales_history_df[date_column] >= one_year_ago]
            # Continue with your grouping and summing logic as before
            sales_history_grouped = sales_history_df_filtered.groupby('Last Ship CPN')['Net'].sum().reset_index()
            # Map these summed 'NET' values to '12 Month CPN Sales' in active_supplier_df
            # Convert the grouped data into a dictionary for easy mapping
            sales_net_mapping = sales_history_grouped.set_index('Last Ship CPN')['Net'].to_dict()
            # Map the summed 'NET' values to 'active_supplier_df' based on 'IPN'
            active_supplier_df['12 Month CPN Sales'] = active_supplier_df['IPN'].map(sales_net_mapping)
            # -----------------------------------------------------------------------------------

            # ------------------------ SND and VPC Logic ----------------------------------------
            '''
            Here we added a cost value that updates the cost column based on if here is a matching product id with 
            the psoft part number and we can match the new cost to that column. 
            '''
            # Check for a matching entry in the SND dataframe
            matching_snd = snd_df[snd_df['Product ID'].astype(str) == str(psoft_part)]
            if not matching_snd.empty:
                matching_record = matching_snd.iloc[0]
                active_supplier_df.at[idx, 'Cost'] = matching_record.get('SND Cost', row['Cost'])  # Update Cost
                active_supplier_df.at[idx, 'Cost Exp Date'] = matching_record.get('SND Exp Date', row['Cost Exp Date'])
                active_supplier_df.at[idx, 'Quote#'] = matching_record.get('SND Quote', row['Quote#'])
                active_supplier_df.at[idx, 'Cost MOQ'] = matching_record.get('SND MOQ', row['Cost MOQ'])

            # Check for a matching entry in the VPC dataframe
            matching_vpc = vpc_df[vpc_df['PART ID'].astype(str) == str(psoft_part)]

            if not matching_vpc.empty:
                matching_record = matching_vpc.iloc[0]
                active_supplier_df.at[idx, 'Cost'] = matching_record.get('VPC Cost', row['Cost'])  # Update Cost
                active_supplier_df.at[idx, 'Cost Exp Date'] = matching_record.get('VPC Exp Date', row['Cost Exp Date'])
                active_supplier_df.at[idx, 'Quote#'] = matching_record.get('VPC Quote', row['Quote#'])
                active_supplier_df.at[idx, 'Cost MOQ'] = matching_record.get('VPC MOQ', row['Cost MOQ'])

        # Note: The .get() method is used for dictionaries; adjust the logic for dataframe access as necessary.
        # This pseudocode assumes 'SND Exp Date', 'SND Quote', 'SND MOQ', 'VPC Exp Date', 'VPC Quote', and 'VPC MOQ'
        # are the correct column names in your SND and VPC dataframes. Adjust as necessary to fit your actual dataframe structures.

        # -----------------------------------------------------------------------------

        # Convert the 'PS Award Exp Date' in active_supplier_df to 'MM-DD-YYYY' format
        active_supplier_df['PS Award Exp Date'] = pd.to_datetime(active_supplier_df['PS Award Exp Date'],
                                                                 errors='coerce').dt.strftime('%m-%d-%Y')
        # --------------------- End of: UPDATE AWARDS DETAILS IN ACTIVE SUPPLIER DATAFRAME--------
        # ---------------------- Contract Change Logic -----------------------------------------------
        active_supplier_df['Contract Change'] = np.where(
            active_supplier_df['LW PRICE'].isna(),  # Check if 'LW PRICE' is NaN or null
            'New Item',  # New Item is going to be populated if the item was not in the prev contract
            np.where(  # Locate where they are going to include the other contingencies
                active_supplier_df['Price'] == active_supplier_df['LW PRICE'],
                'No Change',  # The no change will populate if the price column is identical from one week to the next
                np.where(
                    active_supplier_df['Price'] > active_supplier_df['LW PRICE'],
                    'Price Increase',
                    'Price Decrease'  # Since we've covered all other scenarios, this can be the else condition.
                )
            )
        )
        # --------------------- End of Contract Change Logic --------------------------------------------

        # ------------------- MOQ Match Logic ----------------------------------------------------
        # Merge 'MOQ' from prev_contract_df into active_supplier_df for comparison
        active_supplier_df = pd.merge(
            active_supplier_df,
            prev_contract_df[['IPN', 'MOQ']],
            on='IPN',
            how='left',
            suffixes=('', '_prev')
        )

        # Locate the 'Contract Change' column index
        contract_change_idx = active_supplier_df.columns.get_loc('Contract Change')

        # Add the 'MOQ Match' column next to the 'Contract Change' column
        active_supplier_df.insert(contract_change_idx + 1, 'MOQ Match', '')

        # Populate the 'MOQ Match' column
        active_supplier_df['MOQ Match'] = np.where(
            active_supplier_df['MOQ'] == active_supplier_df['MOQ_prev'], 'Y', 'N'
        )
        active_supplier_df.drop(columns=['MOQ_prev'], inplace=True)
        # ----------------------- End of MOQ ----------------------------------------------------

        # --------------------- VPC Type Logic -----------------------------------
        # This needs to be added before we save the output file you idiot... that's why we needed it here
        # Ensure 'Cost Exp Date' column exists
        if 'Cost MOQ' in active_supplier_df.columns:
            # Find the index of the 'Cost Exp Date' column
            cost_exp_date_index = active_supplier_df.columns.get_loc('Cost MOQ') + 1  # +1 to insert after

            # Step 2: Insert the 'VPC TYPE' column if it doesn't already exist
            if 'VPC TYPE' not in active_supplier_df.columns:
                active_supplier_df.insert(loc=cost_exp_date_index, column='VPC TYPE',
                                          value=np.nan)  # Initialize with NaN
            else:
                # If 'VPC TYPE' already exists but not in the right place, adjust its location.
                active_supplier_df.drop(columns=['VPC TYPE'], inplace=True)
                active_supplier_df.insert(loc=cost_exp_date_index, column='VPC TYPE', value=np.nan)

            # Normalize the 'PART ID' in VPC DataFrame for a more accurate lookup
            vpc_df['PART ID'] = vpc_df['PART ID'].astype(
                str).str.strip().str.upper()  # Assuming case insensitivity

            # Create a mapping from 'PART ID' to 'VPC TYPE' in vpc_df
            vpc_type_mapping = vpc_df.set_index('PART ID')['VPC TYPE'].to_dict()

            # Display some sample mappings
            print("Sample mappings from VPC DataFrame:", list(vpc_type_mapping.items())[:5])

            # Map the 'VPC TYPE' to 'PSoft Part' in active_supplier_df
            active_supplier_df['PSoft Part'] = active_supplier_df['PSoft Part'].astype(
                str).str.strip().str.upper()  # Normalize similarly
            active_supplier_df['VPC TYPE'] = active_supplier_df['PSoft Part'].map(vpc_type_mapping)
        # --------------------- End of VPC Type logic -----------------------------

        # --------------------- Save Output File Logic --------------------------------------------------
        # Ask the user for the output file path
        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Save the output file as",
                                                   initialdir="P:\Partnership_Python_Projects\Creation\test_001")
        # --------------------- End of Save Output File Logic -------------------------------------------

        # --------------------- Saving the desired sheets in final output ------------------------------
        # Write all the DataFrames to the new Excel file in the specified order
        if output_file:
            with (pd.ExcelWriter(output_file, engine='openpyxl') as writer):
                active_supplier_df.to_excel(writer, index=False, sheet_name='Active Supplier Contracts')
                prev_contract_df.to_excel(writer, index=False, sheet_name='Prev Contract')
                lost_items_df.to_excel(writer, index=False, sheet_name='Lost Items')
                awards_df.to_excel(writer, index=False, sheet_name='Awards')
                snd_df.to_excel(writer, index=False, sheet_name='SND')
                vpc_df.to_excel(writer, index=False, sheet_name='VPC')
                backlog_df.to_excel(writer, index=False, sheet_name='Backlog')
                sales_history_df.to_excel(writer, index=False, sheet_name='Sales History')
                running_file_df.to_excel(writer, index=False, sheet_name='Price Increases')

                # Iterate over each sheet in the workbook to freeze top row, wrap text, and turn on filters
                for sheet_name in writer.sheets:
                    sheet = writer.sheets[sheet_name]
                    sheet.freeze_panes = 'A2'  # Freeze the top row
                    sheet.auto_filter.ref = sheet.dimensions  # Turn on filters for the top row only
                    for cell in sheet["1:1"]:
                        cell.alignment = Alignment(wrap_text=True)  # Wrap text for headers

                # Grabbing the workbook and the desired sheet
                workbook = writer.book
                sheet = workbook['Active Supplier Contracts']
                # Freeze the top row
                sheet.freeze_panes = 'A2'
                # Freeze the column H2
                sheet.freeze_panes = "H2"
                # Turn on filters for the top row only
                sheet.auto_filter.ref = sheet.dimensions
                # Wrap text for the first row, makes it look neater
                for cell in sheet["1:1"]:
                    cell.alignment = Alignment(wrap_text=True)
                # --------------------- End of Saving the desired sheets in final output -----------------

                # --------------------- Duplicate Psoft part logic -----------------
                # Applying conditional formatting for duplicates in 'PSoft Part'
                psoft_part_column = get_column_letter(
                    active_supplier_df.columns.get_loc('PSoft Part') + 1)  # +1 as DataFrame columns are 0-indexed

                # Read back the data from the worksheet to find duplicates
                psoft_parts = {}
                for row in range(2, sheet.max_row + 1):  # Skip header row, start from 2
                    cell_value = sheet[f'{psoft_part_column}{row}'].value
                    if cell_value in psoft_parts:
                        psoft_parts[cell_value].append(f'{psoft_part_column}{row}')
                    else:
                        psoft_parts[cell_value] = [f'{psoft_part_column}{row}']

                # Apply light red fill to duplicates
                light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                for cells in psoft_parts.values():
                    if len(cells) > 1:  # More than one occurrence means duplicates
                        for cell in cells:
                            sheet[cell].fill = light_red_fill
                # --------------------- End of Duplicate psoft part logic ----------------

                # --------------------- Additional formatting for specific columns -----------------------------
                # Define the columns for 'Price_x', 'Cost', 'GP%', 'Cost Exp Date', 'Award Date', and 'Last Update Date'
                # Pi Sent date, pi eff date, 12 month CPN Sales, 90 DAY PI - NEW PRICE, PS Award Price
                # LW PRICE, Lw Cost, PS Award Exp Date
                price_x_col, cost_col, gp_col, date_col, award_date_col, last_update_date_col, \
                    pi_sent_date_col, pi_eff_date_col, twelve_month_col, nine_day_pi_col, \
                    ps_award_price_col, lw_price_col, lw_cost_col, ps_award_exp_col, sum_value_col, \
                    avg_value_col, diff_value_col = \
                    None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None

                for col_num, col_cells in enumerate(sheet.columns, start=1):
                    col_val = col_cells[
                        0].value  # header value in current column, 0 because we start at index 0 and header = 1
                    if col_val == 'Price':
                        price_x_col = col_num
                    elif col_val == 'Cost':
                        cost_col = col_num
                    elif col_val == 'GP%':
                        gp_col = col_num
                    elif col_val == 'Cost Exp Date':
                        date_col = col_num
                    elif col_val == 'Award Date':
                        award_date_col = col_num
                    elif col_val == 'Last Update Date':
                        last_update_date_col = col_num
                    elif col_val == 'PI SENT DATE':
                        pi_sent_date_col = col_num
                    elif col_val == 'PI EFF DATE':
                        pi_eff_date_col = col_num
                    elif col_val == '12 Month CPN Sales':
                        twelve_month_col = col_num
                    elif col_val == '90 DAY PI - NEW PRICE':
                        nine_day_pi_col = col_num
                    elif col_val == 'PS Award Price':
                        ps_award_price_col = col_num
                    elif col_val == 'LW PRICE':
                        lw_price_col = col_num
                    elif col_val == 'LW Cost':
                        lw_cost_col = col_num
                    elif col_val == 'PS Award Exp Date':
                        ps_award_exp_col = col_num
                    elif col_val == "SUM":
                        sum_value_col = col_num
                    elif col_val == "AVG":
                        avg_value_col = col_num
                    elif col_val == "DIFF":
                        diff_value_col = col_num

                # Check if all the required columns were found and apply formatting, we look for them using f" get column
                if all([price_x_col, cost_col, gp_col, date_col, award_date_col, last_update_date_col,
                        pi_sent_date_col, pi_eff_date_col, twelve_month_col, nine_day_pi_col, ps_award_price_col,
                        lw_price_col, lw_cost_col, ps_award_exp_col, sum_value_col, avg_value_col, diff_value_col]):
                    for row in range(2, sheet.max_row + 1):  # Assuming row 1 is the header, so we start from row 2
                        gp_cell = f"{get_column_letter(gp_col)}{row}"  # GP % Column formatting
                        price_x_cell = f"{get_column_letter(price_x_col)}{row}"  # Price cell formatting
                        cost_cell = f"{get_column_letter(cost_col)}{row}"  # Cost cell formatting
                        date_cell = f"{get_column_letter(date_col)}{row}"  # Date cell formatting
                        award_date_cell = f"{get_column_letter(award_date_col)}{row}"  # Award Date cell formatting
                        last_update_date_cell = f"{get_column_letter(last_update_date_col)}{row}"  # Last Update cell formatting
                        pi_sent_date_cell = f"{get_column_letter(pi_sent_date_col)}{row}"  # Pi Sent Date cell formatting
                        pi_eff_date_cell = f"{get_column_letter(pi_eff_date_col)}{row}"  # Pi Eff Date cell formatting
                        twelve_month_cell = f"{get_column_letter(twelve_month_col)}{row}"  # 12 Month CPN cell formatting
                        nine_day_pi_cell = f"{get_column_letter(nine_day_pi_col)}{row}"  # 90 Day Price cell formatting
                        ps_award_price_cell = f"{get_column_letter(ps_award_price_col)}{row}"  # PS Awd Price cell formatting
                        lw_price_cell = f"{get_column_letter(lw_price_col)}{row}"  # LW PRICE cell formatting
                        lw_cost_cell = f"{get_column_letter(lw_cost_col)}{row}"  # LW Cost cell formatting
                        ps_award_exp_cell = f"{get_column_letter(ps_award_exp_col)}{row}"  # PS Awd Exp cell formatting
                        sum_cell = f"{get_column_letter(sum_value_col)}{row}"  # Sum Cell formatting
                        avg_cell = f"{get_column_letter(avg_value_col)}{row}"  # Avg cell formatting to numeric value
                        diff_cell = f"{get_column_letter(diff_value_col)}{row}"  # Diff cell formatting to numeric value

                        # Format for cells that we added the first portion of the col value is the column value in the active workbook
                        # For instance lw_cost is LW Cost in our active workbook
                        sheet[price_x_cell].number_format = '$0.0000'  # Formats the Price cells accordingly
                        sheet[gp_cell].number_format = '0.00%'  # GP% as percent
                        sheet[cost_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                        sheet[date_cell].number_format = 'MM/DD/YYYY'  # Date as MM/DD/YYYY
                        sheet[award_date_cell].number_format = 'MM/DD/YYYY'  # Award Date as MM/DD/YYYY
                        sheet[last_update_date_cell].number_format = 'MM/DD/YYYY'  # Last Update Date as MM/DD/YYYY
                        sheet[pi_sent_date_cell].number_format = 'MM/DD/YYYY'  # Format as MM/DD/YYYY
                        sheet[pi_eff_date_cell].number_format = 'MM/DD/YYYY'  # Format as MM/DD/YYYY
                        sheet[twelve_month_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                        sheet[nine_day_pi_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                        sheet[ps_award_price_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                        sheet[lw_price_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                        sheet[lw_cost_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                        sheet[ps_award_exp_cell].number_format = 'MM/DD/YYYY'  # Format as MM/DD/YYYY
                        sheet[sum_cell].number_format = '$0.0000'  # SUM as dollar with four decimal places
                        sheet[avg_cell].number_format = '$0.0000'  # AVG as dollar with four decimal places
                        sheet[diff_cell].number_format = '$0.0000'  # DIFF as dollar with four decimal places

                        # Apply formula to GP%, this is the formula Jess provided for us (price - cost) / price
                        formula = f"=IF({price_x_cell}=0,0,({price_x_cell} - {cost_cell}) / {price_x_cell})"
                        sheet[gp_cell] = formula

                for col_num, col_cells in enumerate(sheet.columns, start=1):
                    if col_cells[0].value in headers_to_color:
                        col_cells[0].fill = PatternFill(start_color=headers_to_color[col_cells[0].value],
                                                        end_color=headers_to_color[col_cells[0].value],
                                                        fill_type="solid")

            # Display a success message in a message box
            messagebox.showinfo("Success", "The output file has been saved as: " + output_file)
            button_to_disable.config(state="disabled")

            process = psutil.Process()
            print(f"Memory usage: {process.memory_info().rss / 1024 ** 2:.2f} MB")  # memory usage in MB

    except Exception as e:
        messagebox.showerror("Error Process was Cancelled", str(e))
        process = psutil.Process()
        print(f"Memory usage at error: {process.memory_info().rss / 1024 ** 2:.2f} MB")
        # Exception is to cause the button to clear ability to re-click once complete
        # Adjusting the rate of how the file is being used and how much memory we are using
