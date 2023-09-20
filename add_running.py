from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def add_running_file_to_workbook():
    print("add_running_file_to_workbook called")

    # Ask the user to select the workbook to which the Running File data should be added
    workbook_file_path = filedialog.askopenfilename(title="Select the workbook to add the Running File data",
                                                    filetypes=[("Excel files", "*.xlsx;*.xls")],
                                                    initialdir="I:\Quotes\Partnership Sales - CM\Creation")
    if not workbook_file_path:
        messagebox.showerror("Error", "Workbook selection cancelled.")
        return

    special_file_name = "Running File - 30 Day Notice Contract Price Increase_Sager - COSTED"
    running_file_path = filedialog.askopenfilename(title="Select {} file".format(special_file_name),
                                                   initialdir="I:\Quotes\Partnership Sales - CM\Creation")
    if not running_file_path:
        messagebox.showerror("Error", "File selection for Running File cancelled.")
        return

    # Read the Running File data into a DataFrame
    running_data = pd.read_excel(running_file_path)

    # Load the selected workbook
    selected_workbook = load_workbook(workbook_file_path)
    running_sheet = selected_workbook.create_sheet(title=special_file_name)

    # Transfer data from DataFrame to the selected workbook
    for r in dataframe_to_rows(running_data, index=False, header=True):
        running_sheet.append(r)

    try:
        selected_workbook.save(workbook_file_path)
        messagebox.showinfo("Success!", "Running File data added successfully.")
    except Exception as e:
        messagebox.showerror("Error", "Failed to add Running File. Error: " + str(e))


# Currently this method is still giving us an error
# import sqlite3
# from tkinter import filedialog, messagebox
# from openpyxl import Workbook
# from openpyxl.reader.excel import load_workbook
# import pandas as pd
# from openpyxl.utils.dataframe import dataframe_to_rows
#
#
# def get_table_names(conn):
#     cursor = conn.cursor()
#     cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
#     tables = cursor.fetchall()
#     return [table[0] for table in tables]
#
#
# def add_running_file_to_workbook():
#     # Get the original workbook path
#     workbook_file_path = filedialog.askopenfilename(title="Select the workbook to add the Running File data",
#                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])
#     if not workbook_file_path:
#         messagebox.showerror("Error", "Workbook selection cancelled.")
#         return
#
#     # Get the running file path
#     special_file_name = "Running File - 30 Day Notice Contract Price Increase_Sager - COSTED"
#     running_file_path = filedialog.askopenfilename(title="Select {} file".format(special_file_name))
#     if not running_file_path:
#         messagebox.showerror("Error", "File selection for Running File cancelled.")
#         return
#
#     # Load the workbook using openpyxl
#     original_workbook = load_workbook(workbook_file_path)
#     sheet_names = original_workbook.sheetnames  # Get the names of all sheets
#
#     # Create a temporary SQLite database
#     conn = sqlite3.connect(":memory:")
#
#     # For each sheet, save data to SQLite
#     for sheet_name in sheet_names:
#         data = pd.read_excel(workbook_file_path, sheet_name=sheet_name)
#         data.to_sql(sheet_name, conn, index=False, if_exists='replace')
#
#     # Load running file data and save to SQLite
#     running_data = pd.read_excel(running_file_path)
#     running_data.to_sql(special_file_name, conn, index=False, if_exists='replace')
#
#     # Create a new workbook
#     new_workbook = Workbook()
#     new_workbook.remove(new_workbook.active)  # Remove the default sheet
#
#     # Populate the new workbook with the data from SQLite
#     for sheet_name in sheet_names + [special_file_name]:
#         sheet_data = pd.read_sql(f"SELECT * FROM '{sheet_name}'", conn)
#         new_sheet = new_workbook.create_sheet(sheet_name)
#         for r in dataframe_to_rows(sheet_data, index=False, header=True):
#             new_sheet.append(r)
#
#     # Save the workbook
#     try:
#         new_workbook.save(workbook_file_path)
#         messagebox.showinfo("Success!", "Running File data added successfully.")
#     except Exception as e:
#         messagebox.showerror("Error", "Failed to add Running File. Error: " + str(e))
#     finally:
#         conn.close()
