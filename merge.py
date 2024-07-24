from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def merge_files_and_create_lost_items(button_to_disable):
    print("merge_files_and_create_lost_items called")
    file_names = ["Active Contract File", "Prev Contract", "Awards", "Backlog", "Sales History", "SND", "VPC", 'Running File - 30 Day Notice Co']
    file_paths = []

    for file_name in file_names:
        file_path = filedialog.askopenfilename(title="Select {} file".format(file_name),
                                               initialdir="P:\\Partnership_Python_Projects\\Creation\\WEEKLY CONTRACT FILES")
        if not file_path:
            messagebox.showerror("Error", "File selection cancelled.")
            return
        file_paths.append(file_path)
    active_award_file_path = file_paths[0]
    active_award_workbook = load_workbook(active_award_file_path)

    # Load active and prev contract dataframes
    active_df = pd.read_excel(active_award_file_path, sheet_name=0, skiprows=1, dtype={'IPN': str})
    prev_contract_file_path = file_paths[1]
    prev_df = pd.read_excel(prev_contract_file_path, sheet_name=0, dtype={'IPN': str})

    # Ensure leading zeros are retained and aligned
    max_ipn_length = max(active_df['IPN'].str.len().max(), prev_df['IPN'].str.len().max())
    active_df['IPN'] = active_df['IPN'].apply(lambda x: x.zfill(max_ipn_length))
    prev_df['IPN'] = prev_df['IPN'].apply(lambda x: x.zfill(max_ipn_length))

    print("Active DataFrame loaded:")
    print(active_df.head())
    print("Previous DataFrame loaded:")
    print(prev_df.head())

    # Always create 'Lost Items' sheet
    if 'Lost Items' in active_award_workbook.sheetnames:
        lost_items_sheet = active_award_workbook['Lost Items']
        active_award_workbook.remove(lost_items_sheet)
    lost_items_sheet = active_award_workbook.create_sheet('Lost Items')

    # Add data to 'Lost Items' sheet only if there are lost items
    lost_items_df = prev_df[~prev_df['IPN'].isin(active_df['IPN'])]
    print("Lost Items DataFrame:")
    print(lost_items_df.head())

    if not lost_items_df.empty:  # check if there are lost items
        for r in dataframe_to_rows(lost_items_df, index=False, header=True):
            lost_items_sheet.append(r)
    else:
        lost_items_sheet.append(list(prev_df.columns))  # append headers only

    # Create a new sheet for the Prev Contract file
    prev_contract_sheet_name = "Prev Contract"
    if prev_contract_sheet_name in active_award_workbook.sheetnames:
        prev_contract_sheet = active_award_workbook[prev_contract_sheet_name]
        active_award_workbook.remove(prev_contract_sheet)
    prev_contract_sheet = active_award_workbook.create_sheet(title=prev_contract_sheet_name)
    for r in dataframe_to_rows(prev_df, index=False, header=True):
        prev_contract_sheet.append(r)

    # Load and merge data from other files
    for file_path, file_name in zip(file_paths[2:], file_names[2:]):  # Adjusted to skip the first two files as before
        data = pd.read_excel(file_path)
        print(f"Data from {file_name} loaded:")
        print(data.head())

        # Ensure leading zeros are retained and aligned in other DataFrames
        if 'IPN' in data.columns:
            data['IPN'] = data['IPN'].apply(lambda x: x.zfill(max_ipn_length))

        # Determine a sheet name based on the file name. If it's the running file, name it "Price Increases"
        sheet_name = "Price Increases" if file_name == "Running File - 30 Day Notice Co" else file_name

        # Create a new sheet for each file and add data to it
        if sheet_name in active_award_workbook.sheetnames:
            new_sheet = active_award_workbook[sheet_name]
            active_award_workbook.remove(new_sheet)
        new_sheet = active_award_workbook.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(data, index=False, header=True):
            new_sheet.append(r)

    try:
        active_award_workbook.save(active_award_file_path)
        messagebox.showinfo("Success!",
                            "Files merged and 'Lost Items' sheet created "
                            "successfully with any missing IPN's from last week that "
                            "are not in the current weeks file.")
        button_to_disable.config(state="disabled")
    except Exception as e:
        messagebox.showerror("Error", str(e))
