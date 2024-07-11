from datetime import datetime, timedelta
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl.styles import PatternFill, Alignment
import numpy as np
from openpyxl.utils import get_column_letter
from colored_headers import headers_to_color
import psutil


def load_data(contract_file):
    dataframes = {
        'active_supplier_df': pd.read_excel(contract_file, sheet_name='Active Supplier Contracts', header=1),
        'prev_contract_df': pd.read_excel(contract_file, sheet_name='Prev Contract', header=0,
                                          dtype={'PSoft Part': str}),
        'lost_items_df': pd.read_excel(contract_file, sheet_name='Lost Items'),
        'awards_df': pd.read_excel(contract_file, sheet_name='Awards'),
        'snd_df': pd.read_excel(contract_file, sheet_name='SND'),
        'vpc_df': pd.read_excel(contract_file, sheet_name='VPC'),
        'backlog_df': pd.read_excel(contract_file, sheet_name='Backlog'),
        'sales_history_df': pd.read_excel(contract_file, sheet_name='Sales History'),
        'running_file_df': pd.read_excel(contract_file, sheet_name='Price Increases')
    }
    return dataframes


def format_ipn(ipn, length=7):
    """ Ensure the IPN has leading zeros to a specific length """
    return ipn.zfill(length)


def prepare_data(active_supplier_df, prev_contract_df):
    if 'LW PRICE' in prev_contract_df.columns:
        prev_contract_df.drop('LW PRICE', axis=1, inplace=True)
    prev_contract_df.rename(columns={'Price': 'LW PRICE', 'MPN': 'LW MPN'}, inplace=True)

    # Format IPNs to ensure they maintain leading zeros
    active_supplier_df['IPN'] = active_supplier_df['IPN'].apply(lambda x: format_ipn(str(x)))
    prev_contract_df['IPN'] = prev_contract_df['IPN'].apply(lambda x: format_ipn(str(x)))

    active_supplier_df = pd.merge(
        active_supplier_df,
        prev_contract_df[['IPN', 'LW MPN']],
        on='IPN',
        how='left'
    ).rename(columns={'LW MPN': 'Prev Contract MPN'})

    return active_supplier_df


def update_awards_details(active_supplier_df, awards_df):
    # Ensure 'IPN' and 'Award CPN' are treated as strings without stripping any characters
    active_supplier_df['IPN'] = active_supplier_df['IPN'].astype(str)
    awards_df['Award CPN'] = awards_df['Award CPN'].astype(str)

    # Print statements to check actual values in both columns
    print("Checking normalized IPNs in active_supplier_df:", active_supplier_df['IPN'].unique())
    print("Checking normalized Award CPNs in awards_df:", awards_df['Award CPN'].unique())

    # Iterate over the active supplier DataFrame
    for idx, row in active_supplier_df.iterrows():
        ipn = row['IPN']
        # Apply matching logic to find matching indices in the awards DataFrame
        matching_indices = awards_df['Award CPN'] == ipn

        print(f"Checking IPN: {ipn}, Found matches: {matching_indices.any()}")  # Debugging output

        if matching_indices.any():
            relevant_awards = awards_df[matching_indices]
            # Convert 'End Date' to datetime and format it
            relevant_awards['End Date'] = pd.to_datetime(relevant_awards['End Date'], errors='coerce').dt.strftime(
                '%m-%d-%Y')

            # Drop rows where 'End Date' conversion resulted in NaT and update DataFrame
            relevant_awards = relevant_awards.dropna(subset=['End Date'])
            if not relevant_awards.empty:
                latest_end_date = relevant_awards['End Date'].max()
                active_supplier_df.at[idx, 'PS Award Exp Date'] = latest_end_date

                # Fetch the latest price entry corresponding to the latest end date
                latest_price_row = relevant_awards[relevant_awards['End Date'] == latest_end_date]
                if pd.notna(latest_price_row['Award Price'].iloc[0]):
                    active_supplier_df.at[idx, 'PS Award Price'] = latest_price_row['Award Price'].iloc[0]

            # Update 'PS Awd Cust ID' if a non-NA 'Award Cust ID' is found
            non_na_cust_ids = relevant_awards['Award Cust ID'].dropna()
            if not non_na_cust_ids.empty:
                active_supplier_df.at[idx, 'PS Awd Cust ID'] = non_na_cust_ids.iloc[0]

        # Convert 'PS Award Price' to numeric for all entries after updates
        active_supplier_df['PS Award Price'] = pd.to_numeric(active_supplier_df['PS Award Price'], errors='coerce')

        # Create a dictionary from 'Award CPN' to 'Award Price' for fast lookup
        award_price_mapping = awards_df.set_index('Award CPN')['Award Price'].dropna().to_dict()

        # Update 'Price Match Award' based on a close match to 'PS Award Price'
        active_supplier_df['Price Match Award'] = active_supplier_df.apply(
            lambda x: 'Y' if np.isclose(award_price_mapping.get(x['IPN'], np.nan), x['PS Award Price'],
                                        atol=1e-5) else 'N',
            axis=1
        )

    return active_supplier_df


def update_corp_award_loaded_status(active_supplier_df, awards_df):
    # Normalize 'Award CPN' values from awards_df for a more accurate lookup
    normalized_award_cpn_set = set(awards_df['Award CPN'].astype(str).str.strip().str.lower())
    # Update 'Corp Award Loaded' based on the presence of a normalized 'IPN' in the set
    active_supplier_df['Corp Awd Loaded'] = active_supplier_df['IPN'].astype(str).str.strip().str.lower().apply(
        lambda x: 'Y' if x in normalized_award_cpn_set else 'N'
    )
    return active_supplier_df


def merge_and_calculate_aggregates(active_supplier_df, prev_contract_df):
    # Ensure IPNs are normalized for mergin

    # Merge additional columns from previous contract data
    active_supplier_df = pd.merge(
        active_supplier_df,
        prev_contract_df[['IPN', 'LW PRICE', 'PSoft Part', 'Contract Change', 'count', 'SUM', 'AVG', 'DIFF',
                          'PSID All Contract Prices Same?', '90 DAY PI - NEW PRICE', 'PI SENT DATE',
                          'DIFF Price Increase',
                          'PI EFF DATE', '12 Month CPN Sales', 'GP%', 'Cost', 'Cost Note', 'Quote#', 'Cost Exp Date',
                          'Cost MOQ', 'Review Note']],
        on='IPN', how='left'
    )

    # Calculate aggregates for price-related metrics
    active_supplier_df['SUM'] = active_supplier_df.groupby('PSoft Part')['Price'].transform('sum')
    active_supplier_df['AVG'] = active_supplier_df['SUM'] / active_supplier_df['count']
    active_supplier_df['DIFF'] = active_supplier_df['AVG'] - active_supplier_df['Price']

    # Calculate the counts for each 'PSoft Part'
    psoft_part_counts = active_supplier_df['PSoft Part'].value_counts()
    active_supplier_df['count'] = active_supplier_df['PSoft Part'].map(psoft_part_counts)

    return active_supplier_df


def update_awards_details(active_supplier_df, awards_df):
    # Normalize the 'IPN' and 'Award CPN' for consistent matching
    active_supplier_df['IPN'] = active_supplier_df['IPN'].astype(str).str.strip()
    awards_df['Award CPN'] = awards_df['Award CPN'].astype(str).str.strip()

    # Define a function for IPN matching considering potential non-numeric values
    def match_ipns(awards_ipn, active_ipn):
        return awards_ipn == active_ipn

    # Iterate over the active supplier DataFrame
    for idx, row in active_supplier_df.iterrows():
        ipn = row['IPN']
        # Apply matching logic to find matching indices in the awards DataFrame
        matching_indices = awards_df['Award CPN'] == ipn

        if matching_indices.any():
            relevant_awards = awards_df.loc[matching_indices]
            # Convert 'End Date' to datetime and format it
            relevant_awards.loc[:, 'End Date'] = pd.to_datetime(relevant_awards['End Date'],
                                                                errors='coerce').dt.strftime('%m-%d-%Y')

            # Drop rows where 'End Date' conversion resulted in NaT and update DataFrame
            relevant_awards = relevant_awards.dropna(subset=['End Date'])
            if not relevant_awards.empty:
                latest_end_date = relevant_awards['End Date'].max()
                active_supplier_df.at[idx, 'PS Award Exp Date'] = latest_end_date

                # Fetch the latest price entry corresponding to the latest end date
                latest_price_row = relevant_awards[relevant_awards['End Date'] == latest_end_date]
                if pd.notna(latest_price_row['Award Price'].iloc[0]):
                    active_supplier_df.at[idx, 'PS Award Price'] = latest_price_row['Award Price'].iloc[0]

            # Update 'PS Awd Cust ID' if a non-NA 'Award Cust ID' is found
            non_na_cust_ids = relevant_awards['Award Cust ID'].dropna()
            if not non_na_cust_ids.empty:
                active_supplier_df.at[idx, 'PS Awd Cust ID'] = non_na_cust_ids.iloc[0]

        # Convert 'PS Award Price' to numeric for all entries after updates
        active_supplier_df['PS Award Price'] = pd.to_numeric(active_supplier_df['PS Award Price'], errors='coerce')

        # Create a dictionary from 'Award CPN' to 'Award Price' for fast lookup
        award_price_mapping = awards_df.set_index('Award CPN')['Award Price'].dropna().to_dict()

        # Update 'Price Match Award' based on a close match to 'PS Award Price'
        active_supplier_df['Price Match Award'] = active_supplier_df.apply(
            lambda x: 'Y' if np.isclose(award_price_mapping.get(x['IPN'], np.nan), x['PS Award Price'],
                                        atol=1e-5) else 'N',
            axis=1
        )

    return active_supplier_df


def update_cost_details(active_supplier_df, prev_contract_df):
    # Normalize IPN columns for matching

    # Merge previous contract details into the active supplier DataFrame
    merged_df = pd.merge(
        active_supplier_df,
        prev_contract_df[['IPN', 'Cost', 'Cost Note', 'Quote#', 'Cost Exp Date', 'Review Note']],
        on='IPN', how='left',
        suffixes=('', '_prev')
    )

    # Update the active supplier DataFrame with merged data
    active_supplier_df['LW Cost'] = merged_df['Cost_prev']
    active_supplier_df['LW Cost Note'] = merged_df['Cost Note_prev']
    active_supplier_df['LW Quote#'] = merged_df['Quote#_prev']
    active_supplier_df['LW Cost Exp Date'] = merged_df['Cost Exp Date_prev']
    active_supplier_df['LW Review Note'] = merged_df['Review Note_prev']

    return active_supplier_df


def map_backlog_values(backlog_df, lost_items_df):
    # Create a mapping from 'Backlog CPN' to 'Backlog Value'
    backlog_value_mapping = backlog_df.set_index('Backlog CPN')['Backlog Value'].to_dict()

    # Apply the mapping to 'IPN' column in lost_items_df to get 'Backlog Value'
    lost_items_df['Backlog Value'] = lost_items_df['IPN'].map(backlog_value_mapping)

    # Format the 'Backlog Value' column as currency with two decimal places
    lost_items_df['Backlog Value'] = lost_items_df['Backlog Value'].apply(
        lambda x: "" if pd.isnull(x) or x == 0 else "${:,.2f}".format(x)
    )

    return lost_items_df


def calculate_12_month_cpn_sales(active_supplier_df, sales_history_df):
    date_column = 'Last Ship Date'  # Replace with your actual date column name
    sales_history_df[date_column] = pd.to_datetime(sales_history_df[date_column], errors='coerce')
    one_year_ago = datetime.now() - timedelta(days=365)
    sales_history_filtered = sales_history_df[sales_history_df[date_column] >= one_year_ago]
    sales_history_grouped = sales_history_filtered.groupby('Last Ship CPN')['Net'].sum().reset_index()
    sales_net_mapping = sales_history_grouped.set_index('Last Ship CPN')['Net'].to_dict()
    active_supplier_df['12 Month CPN Sales'] = active_supplier_df['IPN'].map(sales_net_mapping)
    return active_supplier_df


def update_costs_from_snd_vpc(active_supplier_df, snd_df, vpc_df):
    for idx, row in active_supplier_df.iterrows():
        psoft_part = row['PSoft Part']
        # Update from SND DataFrame
        matching_snd = snd_df[snd_df['Product ID'].astype(str) == str(psoft_part)]
        if not matching_snd.empty:
            active_supplier_df.at[idx, 'Cost'] = matching_snd.iloc[0]['SND Cost']
            active_supplier_df.at[idx, 'Cost Exp Date'] = matching_snd.iloc[0]['SND Exp Date']
            active_supplier_df.at[idx, 'Quote#'] = matching_snd.iloc[0]['SND Quote']
            active_supplier_df.at[idx, 'Cost MOQ'] = matching_snd.iloc[0]['SND MOQ']

        # Update from VPC DataFrame
        matching_vpc = vpc_df[vpc_df['PART ID'].astype(str) == str(psoft_part)]
        if not matching_vpc.empty:
            active_supplier_df.at[idx, 'Cost'] = matching_vpc.iloc[0]['VPC Cost']
            active_supplier_df.at[idx, 'Cost Exp Date'] = matching_vpc.iloc[0]['VPC Exp Date']
            active_supplier_df.at[idx, 'Quote#'] = matching_vpc.iloc[0]['VPC Quote']
            active_supplier_df.at[idx, 'Cost MOQ'] = matching_vpc.iloc[0]['VPC MOQ']

    return active_supplier_df


def format_ps_award_exp_date(active_supplier_df):
    active_supplier_df['PS Award Exp Date'] = pd.to_datetime(
        active_supplier_df['PS Award Exp Date'], errors='coerce'
    ).dt.strftime('%m-%d-%Y')
    return active_supplier_df


def update_contract_change_logic(active_supplier_df):
    active_supplier_df['Contract Change'] = np.where(
        active_supplier_df['LW PRICE'].isna(),  # Check if 'LW PRICE' is NaN or null
        'New Item',  # New Item for entries without a last week price
        np.where(  # Evaluate changes between last week's and this week's prices
            active_supplier_df['Price'] == active_supplier_df['LW PRICE'],
            'No Change',  # No price change
            np.where(
                active_supplier_df['Price'] > active_supplier_df['LW PRICE'],
                'Price Increase',  # Price has increased
                'Price Decrease'  # Price has decreased
            )
        )
    )
    return active_supplier_df


def update_moq_match(active_supplier_df, prev_contract_df):
    # Merge 'MOQ' data from previous contract DataFrame
    active_supplier_df = pd.merge(
        active_supplier_df,
        prev_contract_df[['IPN', 'MOQ']],
        on='IPN',
        how='left',
        suffixes=('', '_prev')
    )

    # Find the appropriate index to insert the 'MOQ Match' column next to 'Contract Change'
    contract_change_idx = active_supplier_df.columns.get_loc('Contract Change')

    # Insert the 'MOQ Match' column next to 'Contract Change'
    active_supplier_df.insert(contract_change_idx + 1, 'MOQ Match', '')

    # Determine matches between current and previous MOQ values
    active_supplier_df['MOQ Match'] = np.where(
        active_supplier_df['MOQ'] == active_supplier_df['MOQ_prev'], 'Y', 'N'
    )

    # Drop the temporary '_prev' column used for comparison
    active_supplier_df.drop(columns=['MOQ_prev'], inplace=True)

    return active_supplier_df


def update_vpc_type(active_supplier_df, vpc_df):
    # Ensure 'Cost MOQ' column exists before adding 'VPC TYPE'
    if 'Cost MOQ' in active_supplier_df.columns:
        cost_exp_date_index = active_supplier_df.columns.get_loc('Cost MOQ') + 1  # +1 to insert after

        # Insert 'VPC TYPE' column if it doesn't already exist
        if 'VPC TYPE' not in active_supplier_df.columns:
            active_supplier_df.insert(loc=cost_exp_date_index, column='VPC TYPE', value=np.nan)
        else:
            # If 'VPC TYPE' already exists but not in the right place, adjust its location.
            active_supplier_df.drop(columns=['VPC TYPE'], inplace=True)
            active_supplier_df.insert(loc=cost_exp_date_index, column='VPC TYPE', value=np.nan)

        # Normalize 'PART ID' in VPC DataFrame for a more accurate lookup
        vpc_df['PART ID'] = vpc_df['PART ID'].astype(str).str.strip().str.upper()

        # Create a mapping from 'PART ID' to 'VPC TYPE' in vpc_df
        vpc_type_mapping = vpc_df.set_index('PART ID')['VPC TYPE'].to_dict()

        # Map the 'VPC TYPE' to 'PSoft Part' in active_supplier_df
        active_supplier_df['PSoft Part'] = active_supplier_df['PSoft Part'].astype(str).str.strip().str.upper()
        active_supplier_df['VPC TYPE'] = active_supplier_df['PSoft Part'].map(vpc_type_mapping)

    return active_supplier_df


def update_from_running_file(active_supplier_df, running_file_df):
    current_year = datetime.now().year

    for idx, row in active_supplier_df.iterrows():
        ipn = row['IPN']

        # Debug: Print the IPN being processed
        print(f"Processing IPN: {ipn}")

        # Search for a matching IPN value in the running file DataFrame
        matching_running_file = running_file_df[running_file_df['Creation Part Number'] == ipn]

        if not matching_running_file.empty:
            # Debug: Print the matching row from running_file_df
            print(f"Found matching row in running_file_df for IPN {ipn}:")
            print(matching_running_file.iloc[0])

            # Update '90 DAY PI - NEW PRICE' if not NaN
            if not pd.isna(matching_running_file.iloc[0]['Unit Price (New)']):
                active_supplier_df.at[idx, '90 DAY PI - NEW PRICE'] = matching_running_file.iloc[0]['Unit Price (New)']

            # For "PI SENT DATE" with the year check
            pi_sent_date_raw = matching_running_file.iloc[0]['PI Sent Date']  # Assuming column name
            pi_sent_date = pd.to_datetime(pi_sent_date_raw, errors='coerce')  # Convert to datetime
            if pd.notnull(pi_sent_date) and pi_sent_date.year == current_year:
                active_supplier_df.at[idx, 'PI SENT DATE'] = pi_sent_date.strftime('%Y-%m-%d')

            # For "PI EFF DATE" with the year check
            pi_eff_date_raw = matching_running_file.iloc[0]['PI Eff Date']  # Assuming column name
            pi_eff_date = pd.to_datetime(pi_eff_date_raw, errors='coerce')  # Convert to datetime
            if pd.notnull(pi_eff_date) and pi_eff_date.year == current_year:
                active_supplier_df.at[idx, 'PI EFF DATE'] = pi_eff_date.strftime('%Y-%m-%d')

            # Check if a match is found for price difference calculation
            active_price = active_supplier_df.loc[idx, 'Price']
            running_price = matching_running_file.iloc[0]['Unit Price (New)']

            # Check if both prices are not NaN and calculate the difference
            if not pd.isna(active_price) and not pd.isna(running_price):
                price_difference = active_price - running_price
                active_supplier_df.at[idx, 'DIFF Price Increase'] = price_difference
        else:
            # Debug: Print a message if no matching row is found
            print(f"No matching row found in running_file_df for IPN {ipn}")

    return active_supplier_df


def save_output_file(active_supplier_df, prev_contract_df, lost_items_df, awards_df, snd_df, vpc_df, backlog_df,
                     sales_history_df, running_file_df):
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Save the output file as",
                                               initialdir="P:\Partnership_Python_Projects\Creation\test_001")
    if output_file:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write DataFrames to the workbook
            data_frames = {
                'Active Supplier Contracts': active_supplier_df,
                'Prev Contract': prev_contract_df,
                'Lost Items': lost_items_df,
                'Awards': awards_df,
                'SND': snd_df,
                'VPC': vpc_df,
                'Backlog': backlog_df,
                'Sales History': sales_history_df,
                'Price Increases': running_file_df
            }
            for sheet_name, data_frame in data_frames.items():
                data_frame.to_excel(writer, index=False, sheet_name=sheet_name)
                sheet = writer.sheets[sheet_name]
                # Apply styling and formatting
                sheet.freeze_panes = 'A2'  # Freeze the top row
                sheet.auto_filter.ref = sheet.dimensions  # Turn on filters for the entire range
                for cell in sheet["1:1"]:  # Apply text wrapping to header row
                    cell.alignment = Alignment(wrap_text=True)

            # Apply conditional formatting for duplicates in 'PSoft Part'
            sheet = writer.sheets['Active Supplier Contracts']
            apply_conditional_formatting(sheet)
            # --------------------- Additional formatting for specific columns -----------------------------
            # Define the columns for 'Price_x', 'Cost', 'GP%', 'Cost Exp Date', 'Award Date', and 'Last Update Date'
            # Pi Sent date, pi eff date, 12 month CPN Sales, 90 DAY PI - NEW PRICE, PS Award Price
            # LW PRICE, Lw Cost, PS Award Exp Date
            price_x_col, cost_col, gp_col, date_col, award_date_col, last_update_date_col, \
                pi_sent_date_col, pi_eff_date_col, twelve_month_col, nine_day_pi_col, \
                ps_award_price_col, lw_price_col, lw_cost_col, ps_award_exp_col, sum_value_col, \
                avg_value_col, diff_value_col = \
                None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None

            for col_num, col_cells in enumerate(sheet.columns, start=1):
                col_val = col_cells[
                    0].value  # header value in current column, 0 because we start at index 0 and header = 1
                if col_val == 'Price':
                    price_x_col = col_num
                elif col_val == 'Cost':
                    cost_col = col_num
                elif col_val == 'GP%':
                    gp_col = col_num
                elif col_val == 'Cost Exp Date':
                    date_col = col_num
                elif col_val == 'Award Date':
                    award_date_col = col_num
                elif col_val == 'Last Update Date':
                    last_update_date_col = col_num
                elif col_val == 'PI SENT DATE':
                    pi_sent_date_col = col_num
                elif col_val == 'PI EFF DATE':
                    pi_eff_date_col = col_num
                elif col_val == '12 Month CPN Sales':
                    twelve_month_col = col_num
                elif col_val == '90 DAY PI - NEW PRICE':
                    nine_day_pi_col = col_num
                elif col_val == 'PS Award Price':
                    ps_award_price_col = col_num
                elif col_val == 'LW PRICE':
                    lw_price_col = col_num
                elif col_val == 'LW Cost':
                    lw_cost_col = col_num
                elif col_val == 'PS Award Exp Date':
                    ps_award_exp_col = col_num
                elif col_val == "SUM":
                    sum_value_col = col_num
                elif col_val == "AVG":
                    avg_value_col = col_num
                elif col_val == "DIFF":
                    diff_value_col = col_num

            # Check if all the required columns were found and apply formatting, we look for them using f" get column
            if all([price_x_col, cost_col, gp_col, date_col, award_date_col, last_update_date_col,
                    pi_sent_date_col, pi_eff_date_col, twelve_month_col, nine_day_pi_col, ps_award_price_col,
                    lw_price_col, lw_cost_col, ps_award_exp_col, sum_value_col, avg_value_col, diff_value_col]):
                for row in range(2, sheet.max_row + 1):  # Assuming row 1 is the header, so we start from row 2
                    gp_cell = f"{get_column_letter(gp_col)}{row}"  # GP % Column formatting
                    price_x_cell = f"{get_column_letter(price_x_col)}{row}"  # Price cell formatting
                    cost_cell = f"{get_column_letter(cost_col)}{row}"  # Cost cell formatting
                    date_cell = f"{get_column_letter(date_col)}{row}"  # Date cell formatting
                    award_date_cell = f"{get_column_letter(award_date_col)}{row}"  # Award Date cell formatting
                    last_update_date_cell = f"{get_column_letter(last_update_date_col)}{row}"  # Last Update cell formatting
                    pi_sent_date_cell = f"{get_column_letter(pi_sent_date_col)}{row}"  # Pi Sent Date cell formatting
                    pi_eff_date_cell = f"{get_column_letter(pi_eff_date_col)}{row}"  # Pi Eff Date cell formatting
                    twelve_month_cell = f"{get_column_letter(twelve_month_col)}{row}"  # 12 Month CPN cell formatting
                    nine_day_pi_cell = f"{get_column_letter(nine_day_pi_col)}{row}"  # 90 Day Price cell formatting
                    ps_award_price_cell = f"{get_column_letter(ps_award_price_col)}{row}"  # PS Awd Price cell formatting
                    lw_price_cell = f"{get_column_letter(lw_price_col)}{row}"  # LW PRICE cell formatting
                    lw_cost_cell = f"{get_column_letter(lw_cost_col)}{row}"  # LW Cost cell formatting
                    ps_award_exp_cell = f"{get_column_letter(ps_award_exp_col)}{row}"  # PS Awd Exp cell formatting
                    sum_cell = f"{get_column_letter(sum_value_col)}{row}"  # Sum Cell formatting
                    avg_cell = f"{get_column_letter(avg_value_col)}{row}"  # Avg cell formatting to numeric value
                    diff_cell = f"{get_column_letter(diff_value_col)}{row}"  # Diff cell formatting to numeric value

                    # Format for cells that we added the first portion of the col value is the column value in the active workbook
                    # For instance lw_cost is LW Cost in our active workbook
                    sheet[price_x_cell].number_format = '$0.0000'  # Formats the Price cells accordingly
                    sheet[gp_cell].number_format = '0.00%'  # GP% as percent
                    sheet[cost_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                    sheet[date_cell].number_format = 'MM/DD/YYYY'  # Date as MM/DD/YYYY
                    sheet[award_date_cell].number_format = 'MM/DD/YYYY'  # Award Date as MM/DD/YYYY
                    sheet[last_update_date_cell].number_format = 'MM/DD/YYYY'  # Last Update Date as MM/DD/YYYY
                    sheet[pi_sent_date_cell].number_format = 'MM/DD/YYYY'  # Format as MM/DD/YYYY
                    sheet[pi_eff_date_cell].number_format = 'MM/DD/YYYY'  # Format as MM/DD/YYYY
                    sheet[twelve_month_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                    sheet[nine_day_pi_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                    sheet[ps_award_price_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                    sheet[lw_price_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                    sheet[lw_cost_cell].number_format = '$0.0000'  # Cost as dollar with four decimal places
                    sheet[ps_award_exp_cell].number_format = 'MM/DD/YYYY'  # Format as MM/DD/YYYY
                    sheet[sum_cell].number_format = '$0.0000'  # SUM as dollar with four decimal places
                    sheet[avg_cell].number_format = '$0.0000'  # AVG as dollar with four decimal places
                    sheet[diff_cell].number_format = '$0.0000'  # DIFF as dollar with four decimal places

                    # Apply formula to GP%, this is the formula Jess provided for us (price - cost) / price
                    formula = f"=IF({price_x_cell}=0,0,({price_x_cell} - {cost_cell}) / {price_x_cell})"
                    sheet[gp_cell] = formula

            for col_num, col_cells in enumerate(sheet.columns, start=1):
                if col_cells[0].value in headers_to_color:
                    col_cells[0].fill = PatternFill(start_color=headers_to_color[col_cells[0].value],
                                                    end_color=headers_to_color[col_cells[0].value],
                                                    fill_type="solid")

            # Display a success message in a message box
        messagebox.showinfo("Success", "The output file has been saved as: " + output_file)
    else:
        messagebox.showinfo("Cancelled", "File save was cancelled.")


def apply_conditional_formatting(sheet):
    psoft_part_column = get_column_letter(sheet.max_column + 1)  # Assume 'PSoft Part' is the last column
    psoft_parts = {}
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet[f'{psoft_part_column}{row}'].value
        if cell_value in psoft_parts:
            psoft_parts[cell_value].append(f'{psoft_part_column}{row}')
        else:
            psoft_parts[cell_value] = [f'{psoft_part_column}{row}']
    light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for cells in psoft_parts.values():
        if len(cells) > 1:
            for cell in cells:
                sheet[cell].fill = light_red_fill


def perform_vlookup(button_to_disable):
    try:
        # Step 1: User selects the contract file
        contract_file = filedialog.askopenfilename(
            title="Select the contract file, where we need a vlookup",
            initialdir="H:\Program_Testing_Exec\Sorting_Creation_Updated\testing_new_logic_5_13_24"
        )
        if not contract_file:
            messagebox.showinfo("Cancelled", "File selection was cancelled.")
            return

        # Step 2: Load data from selected file
        dfs = load_data(contract_file)
        active_supplier_df = dfs['active_supplier_df']
        prev_contract_df = dfs['prev_contract_df']
        lost_items_df = dfs['lost_items_df']
        awards_df = dfs['awards_df']
        snd_df = dfs['snd_df']
        vpc_df = dfs['vpc_df']
        backlog_df = dfs['backlog_df']
        sales_history_df = dfs['sales_history_df']
        running_file_df = dfs['running_file_df']

        # Step 3: Process each DataFrame
        active_supplier_df = prepare_data(active_supplier_df, prev_contract_df)
        active_supplier_df = update_awards_details(active_supplier_df, awards_df)
        active_supplier_df = update_corp_award_loaded_status(active_supplier_df, awards_df)
        active_supplier_df = merge_and_calculate_aggregates(active_supplier_df, prev_contract_df)
        active_supplier_df = update_from_running_file(active_supplier_df, running_file_df)
        active_supplier_df = update_cost_details(active_supplier_df, prev_contract_df)
        lost_items_df = map_backlog_values(backlog_df, lost_items_df)
        active_supplier_df = calculate_12_month_cpn_sales(active_supplier_df, sales_history_df)
        active_supplier_df = update_costs_from_snd_vpc(active_supplier_df, snd_df, vpc_df)
        active_supplier_df = format_ps_award_exp_date(active_supplier_df)
        active_supplier_df = update_contract_change_logic(active_supplier_df)
        active_supplier_df = update_moq_match(active_supplier_df, prev_contract_df)
        active_supplier_df = update_vpc_type(active_supplier_df, vpc_df)

        # Step 4: Save the processed data to a new Excel file
        save_output_file(active_supplier_df, prev_contract_df, lost_items_df, awards_df, snd_df, vpc_df, backlog_df,
                         sales_history_df, running_file_df)

        # Disable the button to prevent multiple submissions
        button_to_disable.config(state="disabled")

        # Log memory usage
        process = psutil.Process()
        print(f"Memory usage: {process.memory_info().rss / 1024 ** 2:.2f} MB")

    except Exception as e:
        messagebox.showerror("Error Process was Cancelled", str(e))
        process = psutil.Process()
        print(f"Memory usage at error: {process.memory_info().rss / 1024 ** 2:.2f} MB")
        button_to_disable.config(state="normal")
